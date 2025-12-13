"""
Export Service
Generates various export formats including Excel summary reports
"""
import pandas as pd
import numpy as np
from io import BytesIO
from typing import Dict, List, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList


class ExportService:
    """Handles all export operations"""
    
    @staticmethod
    def generate_summary_excel(
        dataset_info: Dict,
        df: pd.DataFrame,
        quality_report: Dict,
        variables_info: List[Dict]
    ) -> BytesIO:
        """Generate comprehensive Excel summary report"""
        
        output = BytesIO()
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        green_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        yellow_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============ Sheet 1: Executive Summary ============
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Title
        ws['A1'] = "SAV INSIGHT STUDIO - ANKET ÖZETİ"
        ws['A1'].font = Font(bold=True, size=18, color="2563EB")
        ws.merge_cells('A1:E1')
        
        ws['A3'] = f"Dosya: {dataset_info.get('filename', 'Unknown')}"
        ws['A4'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Key Metrics
        ws['A7'] = "TEMEL METRİKLER"
        ws['A7'].font = Font(bold=True, size=14)
        
        metrics = [
            ("Toplam Katılımcı", quality_report.get("total_participants", 0)),
            ("Toplam Değişken", quality_report.get("total_variables", 0)),
            ("Tamamlanan Yanıt", quality_report.get("complete_responses", 0)),
            ("Kısmi Yanıt", quality_report.get("partial_responses", 0)),
            ("Dropout Oranı", f"%{quality_report.get('dropout_rate', 0):.1f}"),
        ]
        
        for i, (label, value) in enumerate(metrics, start=8):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Quality Scores
        ws['A15'] = "VERİ KALİTESİ SKORLARI"
        ws['A15'].font = Font(bold=True, size=14)
        
        scores = [
            ("Genel Skor", quality_report.get("overall_score", 0)),
            ("Tamlık (Completeness)", quality_report.get("completeness_score", 0)),
            ("Geçerlilik (Validity)", quality_report.get("validity_score", 0)),
            ("Tutarlılık (Consistency)", quality_report.get("consistency_score", 0)),
            ("Transformasyon Skoru", quality_report.get("transformation_score", 0)),
        ]
        
        for i, (label, score) in enumerate(scores, start=16):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = f"%{score:.1f}"
            
            # Color based on score
            if score >= 80:
                ws[f'C{i}'] = "✓ İYİ"
                ws[f'C{i}'].fill = green_fill
            elif score >= 60:
                ws[f'C{i}'] = "⚠ ORTA"
                ws[f'C{i}'].fill = yellow_fill
            else:
                ws[f'C{i}'] = "✗ DÜŞÜK"
                ws[f'C{i}'].fill = red_fill
            ws[f'C{i}'].font = Font(bold=True, color="FFFFFF")
        
        # Digital Twin Readiness
        ws['A23'] = "DİJİTAL İKİZ UYGUNLUĞU"
        ws['A23'].font = Font(bold=True, size=14)
        
        readiness = quality_report.get("digital_twin_readiness", "red")
        ws['A24'] = "Durum:"
        ws['B24'] = readiness.upper()
        if readiness == "green":
            ws['B24'].fill = green_fill
            ws['C24'] = "Veri dijital ikiz için UYGUN"
        elif readiness == "yellow":
            ws['B24'].fill = yellow_fill
            ws['C24'] = "Veri işlenebilir ancak DİKKAT gerekli"
        else:
            ws['B24'].fill = red_fill
            ws['C24'] = "Veri dijital ikiz için UYGUN DEĞİL"
        ws['B24'].font = Font(bold=True, color="FFFFFF")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        
        # ============ Sheet 2: Variable Quality ============
        ws2 = wb.create_sheet("Değişken Kalitesi")
        
        headers = ["Değişken Kodu", "Etiket", "Tamamlanma %", "Durum", "Sorunlar"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        var_quality = quality_report.get("variable_quality", [])
        for row, var in enumerate(var_quality, start=2):
            ws2.cell(row=row, column=1, value=var.get("code", ""))
            ws2.cell(row=row, column=2, value=var.get("label", "")[:50])  # Truncate long labels
            ws2.cell(row=row, column=3, value=f"%{var.get('completeness', 0):.1f}")
            
            status = var.get("status", "red")
            status_cell = ws2.cell(row=row, column=4, value=status.upper())
            if status == "green":
                status_cell.fill = green_fill
            elif status == "yellow":
                status_cell.fill = yellow_fill
            else:
                status_cell.fill = red_fill
            status_cell.font = Font(bold=True, color="FFFFFF")
            
            issues = var.get("issues", [])
            ws2.cell(row=row, column=5, value="; ".join(issues) if issues else "-")
        
        # Adjust widths
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 50
        
        # ============ Sheet 3: Recommendations ============
        ws3 = wb.create_sheet("Öneriler")
        
        ws3['A1'] = "KRİTİK SORUNLAR"
        ws3['A1'].font = Font(bold=True, size=14, color="EF4444")
        
        critical = quality_report.get("critical_issues", [])
        for i, issue in enumerate(critical, start=2):
            ws3[f'A{i}'] = f"❌ {issue}"
            ws3[f'A{i}'].font = Font(color="EF4444")
        
        start_row = len(critical) + 4
        ws3[f'A{start_row}'] = "UYARILAR"
        ws3[f'A{start_row}'].font = Font(bold=True, size=14, color="F59E0B")
        
        warnings = quality_report.get("warnings", [])
        for i, warning in enumerate(warnings, start=start_row+1):
            ws3[f'A{i}'] = f"⚠ {warning}"
            ws3[f'A{i}'].font = Font(color="F59E0B")
        
        start_row = start_row + len(warnings) + 3
        ws3[f'A{start_row}'] = "ÖNERİLER"
        ws3[f'A{start_row}'].font = Font(bold=True, size=14, color="22C55E")
        
        suggestions = quality_report.get("recommendations", [])
        for i, suggestion in enumerate(suggestions, start=start_row+1):
            ws3[f'A{i}'] = f"💡 {suggestion}"
        
        ws3.column_dimensions['A'].width = 100
        
        # ============ Sheet 4: Raw Data Sample ============
        ws4 = wb.create_sheet("Veri Örneği")
        
        sample_df = df.head(100)  # First 100 rows
        
        for col, header in enumerate(sample_df.columns, 1):
            cell = ws4.cell(row=1, column=col, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, row in enumerate(sample_df.values, start=2):
            for col_idx, value in enumerate(row, 1):
                ws4.cell(row=row_idx, column=col_idx, value=str(value) if pd.notna(value) else "")
        
        # ============ Sheet 5: Data Dictionary ============
        ws5 = wb.create_sheet("Veri Sözlüğü")
        
        headers = ["Kod", "Etiket", "Tip", "Ölçek", "Kardinalite", "Yanıt Oranı"]
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, var in enumerate(variables_info, start=2):
            ws5.cell(row=row, column=1, value=var.get("code", ""))
            ws5.cell(row=row, column=2, value=var.get("label", ""))
            ws5.cell(row=row, column=3, value=var.get("type", ""))
            ws5.cell(row=row, column=4, value=var.get("measure", ""))
            ws5.cell(row=row, column=5, value=var.get("cardinality", 0))
            ws5.cell(row=row, column=6, value=f"%{var.get('responseRate', 0):.1f}")
        
        ws5.column_dimensions['A'].width = 20
        ws5.column_dimensions['B'].width = 60
        ws5.column_dimensions['C'].width = 15
        ws5.column_dimensions['D'].width = 15
        ws5.column_dimensions['E'].width = 15
        ws5.column_dimensions['F'].width = 15
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_data_excel(df: pd.DataFrame, variables_info: List[Dict]) -> BytesIO:
        """Generate Excel with raw data and labeled data"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Raw data
            df.to_excel(writer, sheet_name='Ham Veri', index=False)
            
            # Create labeled version
            labeled_df = df.copy()
            for var in variables_info:
                if var.get("valueLabels"):
                    value_map = {vl["value"]: vl["label"] for vl in var["valueLabels"]}
                    if var["code"] in labeled_df.columns:
                        labeled_df[var["code"]] = labeled_df[var["code"]].map(
                            lambda x: value_map.get(x, x) if pd.notna(x) else x
                        )
            
            labeled_df.to_excel(writer, sheet_name='Etiketli Veri', index=False)
        
        output.seek(0)
        return output
    
    @staticmethod
    def generate_json_export(dataset_info: Dict, quality_report: Dict) -> Dict:
        """Generate JSON export with all metadata"""
        return {
            "dataset": {
                "id": dataset_info.get("id"),
                "filename": dataset_info.get("filename"),
                "rows": dataset_info.get("nRows"),
                "columns": dataset_info.get("nCols"),
                "created_at": dataset_info.get("createdAt")
            },
            "quality": quality_report,
            "variables": dataset_info.get("variables", []),
            "exported_at": datetime.now().isoformat()
        }

